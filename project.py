
import openpyxl

wb_obj = openpyxl.load_workbook("cena.xlsx")
sheet_obj = wb_obj.active
cenass = sheet_obj.cell(row=1, column = 2)
tovar = sheet_obj.cell(row=1, column = 1)


wb_obj1 = openpyxl.load_workbook("fact.xlsx")
sheet_obj1 = wb_obj1.active

i=1
n=(sheet_obj.max_row)
while i < n:
    tovar = sheet_obj.cell(row=1, column = 1)
    tovar11 = sheet_obj1.cell(row=i, column = 1)
    cenaff = sheet_obj1.cell(row=i, column = 2)
    cenass = sheet_obj.cell(row=1, column = 2)
    if tovar11.value == tovar.value:
        k = cenaff.value
        l = cenass.value
        d = (k/l-1)*100
        i = i + 1
        
    else:
        i = i + 1
        
        
        

cenaff = sheet_obj1.cell(row=1, column = 2)
tovar1 = sheet_obj1.cell(row=1, column = 1)
print('себестоимотсь товара ' + str(tovar.value) + ' равна ' + str(cenass.value))
print('цена продажи товара ' + str(tovar.value) + ' равна ' + str(cenaff.value))      
print('наценка на товар ' + str(tovar.value) + ' равна ' + str(d) )








